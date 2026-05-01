#!/usr/bin/env python3
"""
Export each sheet of docs/assets/MQGT_SCF_Extended_Periodic_Table_1-184.xlsx to CSV
under docs/data/ for diff-friendly revision control.

Requires: pip install 'openpyxl>=3.1'
"""
from __future__ import annotations

import csv
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "docs", "assets", "MQGT_SCF_Extended_Periodic_Table_1-184.xlsx")
OUT_DIR = os.path.join(ROOT, "docs", "data")


def _safe_name(sheet: str) -> str:
    s = re.sub(r"[^\w\-]+", "_", sheet.strip())
    s = re.sub(r"_+", "_", s).strip("_").lower()
    return s or "sheet"


def main() -> int:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        print("Install openpyxl: pip install 'openpyxl>=3.1'", file=sys.stderr)
        return 1

    if not os.path.isfile(XLSX):
        print("Missing workbook:", XLSX, file=sys.stderr)
        return 1

    os.makedirs(OUT_DIR, exist_ok=True)
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    manifest: list[str] = []

    for name in wb.sheetnames:
        ws = wb[name]
        out_name = f"extended_periodic_table_{_safe_name(name)}.csv"
        out_path = os.path.join(OUT_DIR, out_name)
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                w.writerow(["" if v is None else v for v in row])
        manifest.append(out_name)

    manifest_path = os.path.join(OUT_DIR, "extended_periodic_table_manifest.txt")
    with open(manifest_path, "w", encoding="utf-8") as mf:
        mf.write("\n".join(manifest) + "\n")

    print("Wrote", len(manifest), "CSVs to", OUT_DIR)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
